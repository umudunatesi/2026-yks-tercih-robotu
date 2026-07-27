import hashlib, re, unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterator
from openpyxl import load_workbook
from app.core.regions import region_for_city

SHEETS = {"tablo3": ("on_lisans", 35), "tablo4": ("lisans", 48)}
STATUS_MAP = {"yeni": "Yeni", "dolmadı": "Dolmadı", "dolmadi": "Dolmadı", "yer.olmadı": "Yer.Olmadı", "yer.olmadi": "Yer.Olmadı"}

def clean_text(v: Any) -> str | None:
    if v is None: return None
    text = re.sub(r"\s+", " ", str(v).replace("\n", " ")).strip()
    return text or None

def program_code(v: Any) -> str:
    if v is None: return ""
    if isinstance(v, float) and v.is_integer(): return str(int(v))
    return str(v).strip()

def numeric_or_status(v: Any):
    if v is None: return None, None
    if isinstance(v, (int, float)): return v, None
    text = clean_text(v)
    if not text or text in {"--", "----"}: return None, None
    normalized = text.casefold()
    if normalized in STATUS_MAP: return None, STATUS_MAP[normalized]
    try: return float(text.replace(".", "").replace(",", ".")), None
    except ValueError: return None, text

def int_value(v: Any) -> int | None:
    number, _ = numeric_or_status(v)
    return int(number) if number is not None else None

def fold_tr(text: str) -> str:
    table = str.maketrans("ÇĞİIÖŞÜçğıiöşü", "CGIIOSUcgii osu".replace(" ", ""))
    return "".join(c for c in unicodedata.normalize("NFKD", text.translate(table)) if not unicodedata.combining(c)).casefold()

def is_kktc_national_type(value: Any) -> bool:
    """True only for Excel's 'KKTC U.' (KKTC uyruklu) program type."""
    normalized = re.sub(r"[^a-z0-9]+", "", fold_tr(clean_text(value) or ""))
    return normalized == "kktcu"

def map_row(sheet: str, row: tuple[Any, ...]) -> dict:
    is_lisans = sheet == "tablo4"
    university_type = clean_text(row[2] if is_lisans else row[3])
    rank, rank_status = numeric_or_status(row[18])
    history_start = 40 if is_lisans else 27
    extra = {}
    if is_lisans:
        extra = {"fee_2025": clean_text(row[21]), "meb_scholarship_2026": clean_text(row[22]),
                 "school_top_min_score_2025": numeric_or_status(row[24])[0],
                 "school_top_min_rank_2025": int_value(row[25]),
                 "professors": int_value(row[28]), "associate_professors": int_value(row[29]),
                 "assistant_professors": int_value(row[30]), "research_staff": int_value(row[31]),
                 "tus_tt1": clean_text(row[34]), "tus_tt2": clean_text(row[35]), "tus_ktp": clean_text(row[36]),
                 "dus": clean_text(row[37]), "ab_ayp": clean_text(row[38]), "kpss": clean_text(row[39])}
    else:
        extra = {"school_top_min_score_2025": numeric_or_status(row[21])[0],
                 "school_top_min_rank_2025": int_value(row[22])}
    extra["kktc_national_only"] = is_kktc_national_type(university_type)
    history = []
    for offset, year in enumerate(range(2025, 2017, -1)):
        value, status = numeric_or_status(row[history_start + offset])
        history.append({"year": year, "rank": int(value) if value is not None else None, "status": status})
    return {
        "data_year": 2026, "level": "lisans" if is_lisans else "on_lisans", "program_code": program_code(row[1]),
        "sector": None if is_lisans else clean_text(row[2]), "university_type": university_type,
        "region": clean_text(row[3]) if is_lisans else region_for_city(clean_text(row[4])),
        "city": clean_text(row[4]), "location": clean_text(row[5]),
        "university": clean_text(row[6]) or "", "faculty": clean_text(row[7]), "program": clean_text(row[8]) or "",
        "note": clean_text(row[9]), "fee_status": clean_text(row[10]), "language": clean_text(row[11]),
        "education_type": clean_text(row[12]), "duration": clean_text(row[13]), "score_type": clean_text(row[14]),
        "quota_2026": int_value(row[15]), "min_score_2025": numeric_or_status(row[16])[0],
        "max_score_2025": numeric_or_status(row[17])[0], "min_rank_2025": int(rank) if rank is not None else None,
        "rank_status_2025": rank_status, "special_conditions": clean_text(row[19]),
        "threshold_rank": int_value(row[20]) if is_lisans else None,
        "school_top_quota": int_value(row[23] if is_lisans else row[20]),
        "martyr_veteran_quota": int_value(row[26] if is_lisans else row[23]),
        "women_34_quota": int_value(row[27] if is_lisans else row[24]),
        "tyc": clean_text(row[32] if is_lisans else row[25]), "accreditation": clean_text(row[33] if is_lisans else row[26]),
        "extra": extra, "history": history,
    }

def iter_programs(path: str | Path) -> Iterator[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in SHEETS:
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=3, values_only=True):
                if not any(v is not None and str(v).strip() for v in row): continue
                yield map_row(sheet, row)
    finally:
        wb.close()

def analyze(path: str | Path) -> dict:
    counts = {"tablo3": 0, "tablo4": 0}; codes = set(); duplicates = []
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet, (_, width) in SHEETS.items():
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=3, max_col=width, values_only=True):
            if not any(v is not None and str(v).strip() for v in row): continue
            counts[sheet] += 1; code = program_code(row[1])
            if code in codes: duplicates.append(code)
            codes.add(code)
    wb.close()
    expected = {"tablo3": 9253, "tablo4": 12229}
    return {"counts": counts, "total": sum(counts.values()), "expected": expected,
            "matches_expected": counts == expected, "duplicate_codes": duplicates,
            "sha256": hashlib.sha256(Path(path).read_bytes()).hexdigest()}
