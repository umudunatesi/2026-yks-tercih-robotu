import csv
from html import escape
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)
from app.services.special_conditions import condition_details

HEADERS = [
    "Sıra", "Program Kodu", "Üniversite", "Fakülte/YO", "Program", "Şehir",
    "Puan Türü", "2025 Başarı Sırası", "KPSS", "Kategori", "Not",
]
FONT_DIR = Path(__file__).parents[1] / "reports" / "fonts"
pdfmetrics.registerFont(TTFont("Roboto", FONT_DIR / "Roboto-Regular.ttf"))
pdfmetrics.registerFont(TTFont("Roboto-Bold", FONT_DIR / "Roboto-Bold.ttf"))
pdfmetrics.registerFont(TTFont("Roboto-Italic", FONT_DIR / "Roboto-Italic.ttf"))


def safe_csv(value) -> str:
    text = "" if value is None else str(value)
    return "'" + text if text.startswith(("=", "+", "-", "@", "\t", "\r")) else text


def format_kpss(value) -> str:
    if value is None:
        return "—"
    try:
        return f"{float(value):.3f}"
    except (TypeError, ValueError):
        return str(value)


def rows_for_export(items):
    return [
        [
            item.position, item.program.program_code, item.program.university,
            item.program.faculty, item.program.program, item.program.city,
            item.program.score_type, item.program.min_rank_2025,
            (getattr(item.program, "extra", None) or {}).get("kpss"),
            item.category, item.note,
        ]
        for item in items
    ]


def preference_csv(items) -> bytes:
    stream = StringIO(newline="")
    writer = csv.writer(stream, delimiter=";")
    writer.writerow(HEADERS)
    writer.writerows([[safe_csv(x) for x in row] for row in rows_for_export(items)])
    return ("\ufeff" + stream.getvalue()).encode("utf-8")


def preference_xlsx(items) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tercih Listesi"
    ws.append(HEADERS)
    for row in rows_for_export(items):
        ws.append([safe_csv(x) if isinstance(x, str) else x for x in row])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2855D9")
        cell.alignment = Alignment(horizontal="center")
    widths = [8, 16, 34, 32, 34, 16, 12, 20, 14, 18, 28]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def preference_pdf(
    pref_list, student, items, exam_results=None, history_by_program=None
) -> bytes:
    exam_results = exam_results or []
    history_by_program = history_by_program or {}
    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=landscape(A4), leftMargin=5 * mm, rightMargin=5 * mm,
        topMargin=5 * mm, bottomMargin=5 * mm,
    )
    styles = getSampleStyleSheet()
    styles["Title"].fontName = "Roboto-Bold"
    styles["Title"].fontSize = 13
    styles["Title"].leading = 14
    styles["Title"].spaceAfter = 2
    styles["Normal"].fontName = "Roboto"
    styles["Normal"].fontSize = 5.2
    styles["Normal"].leading = 6
    styles["Italic"].fontName = "Roboto-Italic"
    styles["Italic"].fontSize = 5
    styles["Italic"].leading = 6

    selected_types = sorted({
        item.program.score_type for item in items if item.program.score_type
    })
    result_summary = " | ".join(
        f"{result.score_type}: puan "
        f"{result.score if result.score is not None else '—'}, sıra "
        f"{result.rank if result.rank is not None else '—'}"
        for result in exam_results
    ) or "YKS sonucu girilmemiş"
    program_ranks = [
        item.program.min_rank_2025 for item in items
        if item.program.min_rank_2025 is not None
    ]
    rank_range = (
        f"{min(program_ranks):,} – {max(program_ranks):,}".replace(",", ".")
        if program_ranks else "—"
    )
    story = [
        Paragraph("2026 YKS Tercih Listesi", styles["Title"]),
        Paragraph(
            f"<b>Öğrenci:</b> {escape(student.first_name)} "
            f"{escape(student.last_name)} &nbsp;&nbsp; "
            f"<b>Okul:</b> {escape(student.school or '-')} &nbsp;&nbsp; "
            f"<b>Liste:</b> {pref_list.id} / Sürüm {pref_list.version}",
            styles["Normal"],
        ),
        Paragraph(
            f"<b>Seçilen program türleri:</b> "
            f"{escape(', '.join(selected_types) or '—')} &nbsp;&nbsp; "
            f"<b>Listedeki en düşük–en yüksek başarı sırası:</b> {rank_range}",
            styles["Normal"],
        ),
        Paragraph(
            f"<b>Öğrencinin 2026 YKS sonuçları:</b> {escape(result_summary)}",
            styles["Normal"],
        ),
        Spacer(1, 2 * mm),
    ]

    pdf_headers = [
        "Sıra", "Program Kodu", "Program / Üniversite", "Tür",
        "Ek / Öğrenim", "Süre", "Dil", "Ücret durumu",
        "2025 Sıra", "2024 Sıra", "2023 Sıra",
        "2025 Puan", "2024 Puan", "2023 Puan",
        "2026 Kont.", "2025 Kont.", "2024 Kont.",
        "KPSS", "Özel koşullar", "Akreditasyon",
    ]
    body = []
    for item in items:
        program = item.program
        history = history_by_program.get(getattr(program, "id", None), {})
        rank_values = {
            year: (
                history[year].rank
                if year in history and history[year].rank is not None
                else ((history[year].status or "—") if year in history else "—")
            )
            for year in (2025, 2024, 2023)
        }
        score_2025 = (
            f"{getattr(program, 'min_score_2025', None):.3f}"
            if getattr(program, "min_score_2025", None) is not None else "—"
        )
        quota_2026 = (
            getattr(program, "quota_2026", None)
            if getattr(program, "quota_2026", None) is not None else "—"
        )
        education = getattr(program, "education_type", None) or "—"
        extra_info = (
            f"Açıköğretim / {education}"
            if "açık" in education.casefold() else education
        )
        if (getattr(program, "university_type", "") or "").strip().upper() == "KKTC U.":
            extra_info = f"KKTC Uyruklu / {extra_info}"
        special = getattr(program, "special_conditions", None) or "—"
        if len(special) > 90:
            special = special[:87] + "..."
        kpss_text = format_kpss(
            (getattr(program, "extra", None) or {}).get("kpss")
        )
        program_text = (
            f"<b>{escape(program.program)}</b><br/>"
            f"{escape(program.university)} / {escape(program.city or '—')}"
        )
        body.append([
            str(item.position),
            program.program_code,
            Paragraph(program_text, styles["Normal"]),
            program.score_type or "—",
            Paragraph(escape(extra_info), styles["Normal"]),
            getattr(program, "duration", None) or "—",
            getattr(program, "language", None) or "—",
            Paragraph(
                escape(getattr(program, "fee_status", None) or "—"),
                styles["Normal"],
            ),
            Paragraph(f"<b>{rank_values[2025]}</b>", styles["Normal"]),
            Paragraph(str(rank_values[2024]), styles["Normal"]),
            Paragraph(str(rank_values[2023]), styles["Normal"]),
            Paragraph(f"<b>{score_2025}</b>", styles["Normal"]),
            "—",
            "—",
            str(quota_2026),
            "—",
            "—",
            Paragraph(f"<b>{escape(kpss_text)}</b>", styles["Normal"]),
            Paragraph(escape(special), styles["Normal"]),
            Paragraph(
                escape(getattr(program, "accreditation", None) or "—"),
                styles["Normal"],
            ),
        ])

    table = Table(
        [pdf_headers] + body, repeatRows=1,
        colWidths=[
            6 * mm, 18 * mm, 39 * mm, 8 * mm, 13 * mm, 6 * mm,
            12 * mm, 18 * mm, 13 * mm, 13 * mm, 13 * mm,
            13 * mm, 13 * mm, 13 * mm, 11 * mm, 11 * mm, 11 * mm,
            14 * mm, 24 * mm, 14 * mm,
        ],
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2855D9")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Roboto-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Roboto"),
        ("FONTSIZE", (0, 0), (-1, -1), 5.2),
        ("LEADING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 1.2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.2),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F8FAFC")]),
    ]))
    story += [
        table, Spacer(1, 3 * mm),
        Paragraph(
            "Öğrenci İmza: ____________________ &nbsp;&nbsp; "
            "Veli İmza: ____________________ &nbsp;&nbsp; "
            "Rehber Öğretmen: ____________________",
            styles["Normal"],
        ),
        Spacer(1, 2 * mm),
        Paragraph(
            "Bu sistem yalnızca karar destek amacıyla hazırlanmıştır. "
            "Kesin yerleşme garantisi vermez. Güncel ÖSYM/YÖK kılavuzu ve "
            "özel koşullar mutlaka kontrol edilmelidir.",
            styles["Italic"],
        ),
    ]
    used_conditions = {}
    for item in items:
        for condition in condition_details(
            getattr(item.program, "special_conditions", None)
        ):
            used_conditions[condition["code"]] = condition["description"]
    if used_conditions:
        condition_style = styles["Normal"].clone("Condition")
        condition_style.fontSize = 7
        condition_style.leading = 9
        condition_style.spaceAfter = 2.5 * mm
        story.extend([
            PageBreak(),
            Paragraph("Özel Koşul ve Açıklamalar", styles["Title"]),
            Paragraph(
                "Bu bölüm yalnızca tercih listesindeki programlarda kullanılan "
                "koşulları içerir. Kaynak: 2026-YKS Yükseköğretim Programları "
                "ve Kontenjanları Kılavuzu, basılı sayfalar 548-550 ve 552-567.",
                styles["Italic"],
            ),
            Spacer(1, 3 * mm),
        ])
        for code in sorted(used_conditions, key=int):
            story.append(Paragraph(
                f"<b>Bk. {escape(code)}</b> - "
                f"{escape(used_conditions[code])}",
                condition_style,
            ))
    doc.build(story)
    return output.getvalue()
