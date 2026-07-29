import argparse
import json
import re
import unicodedata
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


SHEETS = {
    "tablo3": {
        "width": 35,
        "official_width": 13,
        "mappings": {
            3: 14, 4: 15, 5: 20, 6: 23, 7: 24,
            8: 19, 9: 18, 10: 16, 11: 25, 12: 26,
        },
        "history_2025": 27,
        "identity": [2, 3, 4, 5, 6, 7],
    },
    "tablo4": {
        "width": 48,
        "official_width": 24,
        "mappings": {
            3: 14, 4: 15, 5: 23, 6: 22, 7: 26, 8: 27,
            9: 19, 10: 18, 11: 16, 12: 28, 13: 29, 14: 30,
            15: 31, 16: 32, 17: 33, 18: 34, 19: 35, 20: 36,
            21: 37, 22: 38, 23: 39,
        },
        "history_2025": 40,
        "identity": [2, 3, 4, 5, 6, 7],
    },
}


def code_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def is_university_header(text: str) -> bool:
    return bool(re.search(r"\((Devlet|Vakıf|KKTC|Yurt Dışı)", text))


def fold_text(value: str) -> str:
    return "".join(
        character for character in unicodedata.normalize(
            "NFKD", value.casefold()
        )
        if not unicodedata.combining(character)
    )


def official_records(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    university = None
    university_accreditation = None
    faculty = None
    records = []
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=4, values_only=True), 4
    ):
        code = code_text(row[0])
        text = str(row[1] or "").strip()
        if not code:
            if not text:
                continue
            if is_university_header(text):
                university = text
                university_accreditation = row[
                    12 if sheet.max_column == 13 else 17
                ]
                faculty = None
            else:
                faculty = text
            continue
        if not code.isdigit():
            continue
        records.append({
            "code": code,
            "row_number": row_number,
            "university_header": university,
            "university_accreditation": university_accreditation,
            "faculty_header": faculty,
            "values": tuple(row),
        })
    workbook.close()
    return records


def parse_program_name(raw_name: str) -> dict:
    name = re.sub(r"\s+", " ", raw_name).strip()
    result = {
        "program": name,
        "language": None,
        "fee_status": None,
        "education_type": None,
        "kktc_national_only": False,
    }
    markers = re.findall(r"\(([^()]*)\)", name)
    removable = []
    for marker in markers:
        folded = fold_text(marker)
        if folded == "kktc uyruklu":
            result["kktc_national_only"] = True
            removable.append(marker)
        elif folded in {
            "ingilizce", "almanca", "arapça", "fransızca", "rusça",
        }:
            result["language"] = marker
            removable.append(marker)
        elif folded in {"ücretli", "burslu"} or "indirimli" in folded:
            result["fee_status"] = marker
            removable.append(marker)
        elif folded in {"açıköğretim", "uzaktan öğretim", "iö"}:
            result["education_type"] = marker
            removable.append(marker)
    for marker in removable:
        name = name.replace(f"({marker})", "")
    result["program"] = re.sub(r"\s+", " ", name).strip()
    return result


def normalized_value(value, target_index: int, sheet_name: str):
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
    integer_targets = {
        "tablo3": {15, 18, 20, 23, 24, 27},
        "tablo4": {15, 18, 23, 26, 27, 28, 29, 30, 31, 40},
    }[sheet_name]
    float_targets = {16}
    if target_index in {18, 27, 40} and value == "...":
        return "Yeni"
    if target_index in integer_targets and value not in {"...", "----", "--"}:
        try:
            return int(float(str(value).replace(",", ".")))
        except (TypeError, ValueError):
            return value
    if target_index in float_targets and value not in {"...", "----", "--"}:
        try:
            return float(str(value).replace(",", "."))
        except (TypeError, ValueError):
            return value
    return value


def latest_value(
    record: dict, source_index: int, target_index: int, sheet_name: str,
    for_new: bool = False,
):
    value = record["values"][source_index]
    accreditation_target = (
        target_index == (26 if sheet_name == "tablo3" else 33)
    )
    if for_new and accreditation_target and (
        value is None or str(value).strip() == ""
    ):
        value = record["university_accreditation"]
    return normalized_value(value, target_index, sheet_name)


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def set_latest_values(
    target: list, record: dict, config: dict, sheet_name: str
) -> None:
    for source_index, target_index in config["mappings"].items():
        target[target_index] = latest_value(
            record, source_index, target_index, sheet_name, for_new=True
        )
    history_index = config["history_2025"]
    target[history_index] = normalized_value(
        record["values"][9 if config["official_width"] == 13 else 10],
        history_index,
        sheet_name,
    )
    if target[18] is None:
        target[18] = "Yeni"
        target[history_index] = "Yeni"


def merge_sheet(workbook, sheet_name: str, official_path: Path) -> dict:
    config = SHEETS[sheet_name]
    sheet = workbook[sheet_name]
    records = official_records(official_path)
    if len({record["code"] for record in records}) != len(records):
        raise ValueError(f"{sheet_name}: yinelenen program kodu var")

    existing = {}
    for row_number in range(3, sheet.max_row + 1):
        code = code_text(sheet.cell(row_number, 2).value)
        if code:
            existing[code] = row_number

    official_codes = {record["code"] for record in records}
    added = sorted(official_codes - set(existing))
    removed = sorted(set(existing) - official_codes)
    groups = {}
    for record in records:
        groups.setdefault(record["university_header"], []).append(record)

    changed_fields = 0
    for record in records:
        code = record["code"]
        if code not in existing:
            continue
        row_number = existing[code]
        for source_index, target_index in config["mappings"].items():
            accreditation_target = (
                target_index == (26 if sheet_name == "tablo3" else 33)
            )
            raw_value = record["values"][source_index]
            if accreditation_target and (
                raw_value is None or str(raw_value).strip() == ""
            ):
                continue
            new_value = latest_value(
                record, source_index, target_index, sheet_name
            )
            if target_index == 18 and new_value is None:
                continue
            cell = sheet.cell(row_number, target_index + 1)
            if cell.value != new_value:
                changed_fields += 1
                cell.value = new_value
        history_index = config["history_2025"]
        history_value = normalized_value(
            record["values"][9 if config["official_width"] == 13 else 10],
            history_index,
            sheet_name,
        )
        history_cell = sheet.cell(row_number, config["history_2025"] + 1)
        if history_value is None:
            continue
        if history_cell.value != history_value:
            changed_fields += 1
            history_cell.value = history_value

    max_order = max(
        int(sheet.cell(row_number, 1).value or 0)
        for row_number in existing.values()
    )
    added_details = []
    for code in added:
        record = next(item for item in records if item["code"] == code)
        university_records = groups[record["university_header"]]
        same_faculty = [
            item for item in university_records
            if item["faculty_header"] == record["faculty_header"]
            and item["code"] in existing
        ]
        sibling = (same_faculty or [
            item for item in university_records if item["code"] in existing
        ])
        if not sibling:
            raise ValueError(
                f"{sheet_name}/{code}: üniversite bilgisi için eşleşen kayıt yok"
            )
        base_row = existing[sibling[0]["code"]]
        target_row = sheet.max_row + 1
        for column in range(1, config["width"] + 1):
            copy_cell_style(
                sheet.cell(base_row, column),
                sheet.cell(target_row, column),
            )

        values = [None] * config["width"]
        max_order += 1
        values[0] = max_order
        values[1] = int(code)
        for target_index in config["identity"]:
            values[target_index] = sheet.cell(
                base_row, target_index + 1
            ).value

        if not same_faculty and record["faculty_header"]:
            values[7] = record["faculty_header"]
        parsed = parse_program_name(str(record["values"][1] or ""))
        values[8] = parsed["program"]
        values[10] = parsed["fee_status"]
        values[11] = parsed["language"]
        values[12] = parsed["education_type"]
        values[13] = normalized_value(
            record["values"][2], 13, sheet_name
        )
        if sheet_name == "tablo3" and parsed["kktc_national_only"]:
            values[3] = "KKTC U."
        if sheet_name == "tablo4" and parsed["kktc_national_only"]:
            values[2] = "KKTC U."
        if sheet_name == "tablo4":
            folded_program = fold_text(parsed["program"])
            threshold_rules = [
                ("tip", 50_000), ("dis hekimligi", 80_000),
                ("hukuk", 100_000), ("eczacilik", 100_000),
                ("mimarlik", 250_000), ("muhendisligi", 300_000),
                ("ogretmenligi", 300_000),
            ]
            values[20] = next((
                limit for keyword, limit in threshold_rules
                if keyword in folded_program
            ), None)
        set_latest_values(values, record, config, sheet_name)

        for index, value in enumerate(values, 1):
            sheet.cell(target_row, index).value = value
        added_details.append({
            "code": code,
            "university": values[6],
            "faculty": values[7],
            "program": values[8],
            "score_type": values[14],
        })

    for row_number in sorted(
        (existing[code] for code in removed), reverse=True
    ):
        sheet.delete_rows(row_number, 1)

    final_codes = {
        code_text(sheet.cell(row_number, 2).value)
        for row_number in range(3, sheet.max_row + 1)
        if code_text(sheet.cell(row_number, 2).value)
    }
    if final_codes != official_codes:
        raise ValueError(f"{sheet_name}: çıktı kodları ÖSYM kaynağıyla eşleşmiyor")
    return {
        "sheet": sheet_name,
        "official_count": len(records),
        "final_count": len(final_codes),
        "added": added_details,
        "removed": removed,
        "changed_fields": changed_fields,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base", type=Path, required=True)
    parser.add_argument("--tablo3", type=Path, required=True)
    parser.add_argument("--tablo4", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--report", type=Path)
    args = parser.parse_args()

    workbook = load_workbook(args.base)
    report = {
        "tablo3": merge_sheet(workbook, "tablo3", args.tablo3),
        "tablo4": merge_sheet(workbook, "tablo4", args.tablo4),
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    workbook.close()
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(
            json.dumps(report, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
