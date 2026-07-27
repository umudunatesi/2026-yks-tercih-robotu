import argparse
import json
import sqlite3
from pathlib import Path

from openpyxl import load_workbook


def program_code(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def number(value, integer=False):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value) if integer else float(value)
    text = str(value).strip()
    if not text or text in {"--", "----"}:
        return None
    try:
        parsed = float(text.replace(".", "").replace(",", "."))
        return int(parsed) if integer else parsed
    except ValueError:
        return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("database", type=Path)
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    values = {}
    try:
        for sheet_name, score_index, rank_index in (
            ("tablo3", 21, 22),
            ("tablo4", 24, 25),
        ):
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=3, max_col=26, values_only=True):
                code = program_code(row[1])
                if code:
                    values[code] = (
                        number(row[score_index]),
                        number(row[rank_index], integer=True),
                    )
    finally:
        workbook.close()

    connection = sqlite3.connect(args.database)
    try:
        rows = connection.execute(
            "SELECT program_code, extra FROM programs WHERE data_year = 2026"
        ).fetchall()
        updates = []
        score_count = 0
        rank_count = 0
        for code, raw_extra in rows:
            if code not in values:
                continue
            score, rank = values[code]
            try:
                extra = json.loads(raw_extra) if raw_extra else {}
            except (TypeError, json.JSONDecodeError):
                extra = {}
            extra["school_top_min_score_2025"] = score
            extra["school_top_min_rank_2025"] = rank
            score_count += score is not None
            rank_count += rank is not None
            updates.append((json.dumps(extra, ensure_ascii=False), code))

        with connection:
            connection.executemany(
                "UPDATE programs SET extra = ? "
                "WHERE data_year = 2026 AND program_code = ?",
                updates,
            )
        print(
            f"updated={len(updates)} "
            f"school_top_scores={score_count} school_top_ranks={rank_count}"
        )
    finally:
        connection.close()


if __name__ == "__main__":
    main()
